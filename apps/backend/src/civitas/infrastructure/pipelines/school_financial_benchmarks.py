from __future__ import annotations

import hashlib
import json
import re
import urllib.request
from collections.abc import Callable, Iterable, Sequence
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import cast
from urllib.parse import urlsplit

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text
from sqlalchemy.engine import Engine

from .base import PipelineRunContext, PipelineSource, StageResult, chunked
from .contracts import school_financial_benchmarks as finance_contract

BRONZE_MANIFEST_FILE_NAME = "school_financial_benchmarks.manifest.json"
DEFAULT_WORKBOOK_URLS: tuple[str, ...] = (
    "https://financial-benchmarking-and-insights-tool.education.gov.uk/files/"
    "AAR_2023-24_download.xlsx",
)
_REQUIRED_WORKSHEET_HEADERS = frozenset({"URN", "School Name"})
_ACADEMIC_YEAR_PATTERN = re.compile(r"(?P<start>20\d{2})[-_](?P<end>\d{2}|\d{4})")


@dataclass(frozen=True)
class _ManifestAsset:
    source_url: str
    bronze_file_name: str
    downloaded_at: str
    sha256: str
    academic_year: str
    sheet_names: tuple[str, ...]
    sheet_row_counts: dict[str, int]


@dataclass(frozen=True)
class _StagedRow:
    urn: str
    academic_year: str
    finance_source: str
    school_laestab: str | None
    school_name: str
    trust_uid: str | None
    trust_name: str | None
    phase: str | None
    overall_phase: str | None
    admissions_policy: str | None
    urban_rural: str | None
    pupils_fte: float | None
    teachers_fte: float | None
    fsm_pct: float | None
    ehcp_pct: float | None
    sen_support_pct: float | None
    eal_pct: float | None
    total_grant_funding_gbp: float | None
    total_self_generated_funding_gbp: float | None
    total_income_gbp: float | None
    teaching_staff_costs_gbp: float | None
    supply_teaching_staff_costs_gbp: float | None
    education_support_staff_costs_gbp: float | None
    other_staff_costs_gbp: float | None
    total_staff_costs_gbp: float | None
    maintenance_improvement_costs_gbp: float | None
    premises_costs_gbp: float | None
    educational_supplies_costs_gbp: float | None
    bought_in_professional_services_costs_gbp: float | None
    catering_costs_gbp: float | None
    total_expenditure_gbp: float | None
    revenue_reserve_gbp: float | None
    in_year_balance_gbp: float | None
    source_file_url: str
    source_updated_at_utc: datetime


class SchoolFinancialBenchmarksPipeline:
    source = PipelineSource.SCHOOL_FINANCIAL_BENCHMARKS

    def __init__(
        self,
        engine: Engine | None,
        *,
        workbook_urls: tuple[str, ...] = DEFAULT_WORKBOOK_URLS,
        fetcher: Callable[[str], bytes] | None = None,
    ) -> None:
        normalized_urls = tuple(dict.fromkeys(_normalize_source_url(url) for url in workbook_urls))
        if len(normalized_urls) == 0:
            raise ValueError("At least one school financial benchmarks workbook URL is required.")
        self._engine = engine
        self._workbook_urls = normalized_urls
        self._fetch_workbook = fetcher or _download_workbook

    def download(self, context: PipelineRunContext) -> int:
        context.bronze_source_path.mkdir(parents=True, exist_ok=True)
        manifest_path = context.bronze_source_path / BRONZE_MANIFEST_FILE_NAME
        cached_assets = _load_manifest_assets(manifest_path)
        if cached_assets and _manifest_assets_exist(
            cached_assets,
            bronze_source_path=context.bronze_source_path,
        ):
            return sum(asset.sheet_row_counts.get("Academies", 0) for asset in cached_assets)

        _clear_existing_bronze_files(context.bronze_source_path)
        assets: list[_ManifestAsset] = []
        downloaded_rows = 0
        for source_url in self._workbook_urls:
            workbook_bytes = self._fetch_workbook(source_url)
            bronze_file_name = _build_bronze_file_name(source_url)
            bronze_file_path = context.bronze_source_path / bronze_file_name
            bronze_file_path.write_bytes(workbook_bytes)

            workbook = load_workbook(
                filename=BytesIO(workbook_bytes),
                read_only=True,
                data_only=True,
            )
            try:
                sheet_names = tuple(workbook.sheetnames)
                sheet_row_counts = {
                    sheet_name: _worksheet_row_count(workbook[sheet_name])
                    for sheet_name in workbook.sheetnames
                }
            finally:
                workbook.close()

            asset = _ManifestAsset(
                source_url=source_url,
                bronze_file_name=bronze_file_name,
                downloaded_at=context.started_at.isoformat(),
                sha256=_sha256_bytes(workbook_bytes),
                academic_year=_parse_academic_year(source_url),
                sheet_names=sheet_names,
                sheet_row_counts=sheet_row_counts,
            )
            assets.append(asset)
            downloaded_rows += asset.sheet_row_counts.get("Academies", 0)

        manifest_path.write_text(
            json.dumps(
                {
                    "downloaded_at": context.started_at.isoformat(),
                    "normalization_contract_version": finance_contract.CONTRACT_VERSION,
                    "assets": [
                        {
                            "source_url": asset.source_url,
                            "bronze_file_name": asset.bronze_file_name,
                            "downloaded_at": asset.downloaded_at,
                            "sha256": asset.sha256,
                            "academic_year": asset.academic_year,
                            "sheet_names": list(asset.sheet_names),
                            "sheet_row_counts": asset.sheet_row_counts,
                        }
                        for asset in assets
                    ],
                },
                indent=2,
                sort_keys=True,
            ),
            encoding="utf-8",
        )
        return downloaded_rows

    def stage(self, context: PipelineRunContext) -> StageResult:
        if self._engine is None:
            raise ValueError("Pipeline engine is required for stage.")

        manifest_assets = _load_manifest_assets(
            context.bronze_source_path / BRONZE_MANIFEST_FILE_NAME
        )
        if len(manifest_assets) == 0:
            raise FileNotFoundError(
                "School financial benchmarks manifest is missing; run download first."
            )

        rows_by_key: dict[tuple[str, str], _StagedRow] = {}
        rejected_rows: list[tuple[str, dict[str, object]]] = []
        for asset in manifest_assets:
            workbook_path = context.bronze_source_path / asset.bronze_file_name
            workbook = load_workbook(
                filename=workbook_path,
                read_only=True,
                data_only=True,
            )
            try:
                if "Academies" not in workbook.sheetnames:
                    raise ValueError("Workbook is missing required 'Academies' worksheet.")
                for raw_row in _worksheet_rows(workbook["Academies"]):
                    normalized_row, rejection = finance_contract.normalize_row(
                        raw_row,
                        academic_year=asset.academic_year,
                        source_file_url=asset.source_url,
                    )
                    if normalized_row is None:
                        if rejection is not None:
                            rejected_rows.append((rejection, raw_row))
                        continue
                    staged = _StagedRow(
                        urn=normalized_row["urn"],
                        academic_year=normalized_row["academic_year"],
                        finance_source=normalized_row["finance_source"],
                        school_laestab=normalized_row["school_laestab"],
                        school_name=normalized_row["school_name"],
                        trust_uid=normalized_row["trust_uid"],
                        trust_name=normalized_row["trust_name"],
                        phase=normalized_row["phase"],
                        overall_phase=normalized_row["overall_phase"],
                        admissions_policy=normalized_row["admissions_policy"],
                        urban_rural=normalized_row["urban_rural"],
                        pupils_fte=normalized_row["pupils_fte"],
                        teachers_fte=normalized_row["teachers_fte"],
                        fsm_pct=normalized_row["fsm_pct"],
                        ehcp_pct=normalized_row["ehcp_pct"],
                        sen_support_pct=normalized_row["sen_support_pct"],
                        eal_pct=normalized_row["eal_pct"],
                        total_grant_funding_gbp=normalized_row["total_grant_funding_gbp"],
                        total_self_generated_funding_gbp=normalized_row[
                            "total_self_generated_funding_gbp"
                        ],
                        total_income_gbp=normalized_row["total_income_gbp"],
                        teaching_staff_costs_gbp=normalized_row["teaching_staff_costs_gbp"],
                        supply_teaching_staff_costs_gbp=normalized_row[
                            "supply_teaching_staff_costs_gbp"
                        ],
                        education_support_staff_costs_gbp=normalized_row[
                            "education_support_staff_costs_gbp"
                        ],
                        other_staff_costs_gbp=normalized_row["other_staff_costs_gbp"],
                        total_staff_costs_gbp=normalized_row["total_staff_costs_gbp"],
                        maintenance_improvement_costs_gbp=normalized_row[
                            "maintenance_improvement_costs_gbp"
                        ],
                        premises_costs_gbp=normalized_row["premises_costs_gbp"],
                        educational_supplies_costs_gbp=normalized_row[
                            "educational_supplies_costs_gbp"
                        ],
                        bought_in_professional_services_costs_gbp=normalized_row[
                            "bought_in_professional_services_costs_gbp"
                        ],
                        catering_costs_gbp=normalized_row["catering_costs_gbp"],
                        total_expenditure_gbp=normalized_row["total_expenditure_gbp"],
                        revenue_reserve_gbp=normalized_row["revenue_reserve_gbp"],
                        in_year_balance_gbp=normalized_row["in_year_balance_gbp"],
                        source_file_url=normalized_row["source_file_url"],
                        source_updated_at_utc=_parse_datetime(asset.downloaded_at),
                    )
                    rows_by_key[(staged.urn, staged.academic_year)] = staged
            finally:
                workbook.close()

        staged_rows = sorted(
            rows_by_key.values(),
            key=lambda row: (row.urn, _academic_year_sort_key(row.academic_year)),
        )
        staging_table_name = self._staging_table_name(context)

        with self._engine.begin() as connection:
            connection.execute(text("CREATE SCHEMA IF NOT EXISTS staging"))
            connection.execute(text(f"DROP TABLE IF EXISTS staging.{staging_table_name}"))
            connection.execute(
                text(
                    f"""
                    CREATE TABLE staging.{staging_table_name} (
                        urn text NOT NULL,
                        academic_year text NOT NULL,
                        finance_source text NOT NULL,
                        school_laestab text NULL,
                        school_name text NOT NULL,
                        trust_uid text NULL,
                        trust_name text NULL,
                        phase text NULL,
                        overall_phase text NULL,
                        admissions_policy text NULL,
                        urban_rural text NULL,
                        pupils_fte double precision NULL,
                        teachers_fte double precision NULL,
                        fsm_pct double precision NULL,
                        ehcp_pct double precision NULL,
                        sen_support_pct double precision NULL,
                        eal_pct double precision NULL,
                        total_grant_funding_gbp double precision NULL,
                        total_self_generated_funding_gbp double precision NULL,
                        total_income_gbp double precision NULL,
                        teaching_staff_costs_gbp double precision NULL,
                        supply_teaching_staff_costs_gbp double precision NULL,
                        education_support_staff_costs_gbp double precision NULL,
                        other_staff_costs_gbp double precision NULL,
                        total_staff_costs_gbp double precision NULL,
                        maintenance_improvement_costs_gbp double precision NULL,
                        premises_costs_gbp double precision NULL,
                        educational_supplies_costs_gbp double precision NULL,
                        bought_in_professional_services_costs_gbp double precision NULL,
                        catering_costs_gbp double precision NULL,
                        total_expenditure_gbp double precision NULL,
                        revenue_reserve_gbp double precision NULL,
                        in_year_balance_gbp double precision NULL,
                        source_file_url text NOT NULL,
                        source_updated_at_utc timestamptz NOT NULL,
                        PRIMARY KEY (urn, academic_year)
                    )
                    """
                )
            )
            if staged_rows:
                insert = text(
                    f"""
                    INSERT INTO staging.{staging_table_name} (
                        urn,
                        academic_year,
                        finance_source,
                        school_laestab,
                        school_name,
                        trust_uid,
                        trust_name,
                        phase,
                        overall_phase,
                        admissions_policy,
                        urban_rural,
                        pupils_fte,
                        teachers_fte,
                        fsm_pct,
                        ehcp_pct,
                        sen_support_pct,
                        eal_pct,
                        total_grant_funding_gbp,
                        total_self_generated_funding_gbp,
                        total_income_gbp,
                        teaching_staff_costs_gbp,
                        supply_teaching_staff_costs_gbp,
                        education_support_staff_costs_gbp,
                        other_staff_costs_gbp,
                        total_staff_costs_gbp,
                        maintenance_improvement_costs_gbp,
                        premises_costs_gbp,
                        educational_supplies_costs_gbp,
                        bought_in_professional_services_costs_gbp,
                        catering_costs_gbp,
                        total_expenditure_gbp,
                        revenue_reserve_gbp,
                        in_year_balance_gbp,
                        source_file_url,
                        source_updated_at_utc
                    ) VALUES (
                        :urn,
                        :academic_year,
                        :finance_source,
                        :school_laestab,
                        :school_name,
                        :trust_uid,
                        :trust_name,
                        :phase,
                        :overall_phase,
                        :admissions_policy,
                        :urban_rural,
                        :pupils_fte,
                        :teachers_fte,
                        :fsm_pct,
                        :ehcp_pct,
                        :sen_support_pct,
                        :eal_pct,
                        :total_grant_funding_gbp,
                        :total_self_generated_funding_gbp,
                        :total_income_gbp,
                        :teaching_staff_costs_gbp,
                        :supply_teaching_staff_costs_gbp,
                        :education_support_staff_costs_gbp,
                        :other_staff_costs_gbp,
                        :total_staff_costs_gbp,
                        :maintenance_improvement_costs_gbp,
                        :premises_costs_gbp,
                        :educational_supplies_costs_gbp,
                        :bought_in_professional_services_costs_gbp,
                        :catering_costs_gbp,
                        :total_expenditure_gbp,
                        :revenue_reserve_gbp,
                        :in_year_balance_gbp,
                        :source_file_url,
                        :source_updated_at_utc
                    )
                    ON CONFLICT (urn, academic_year) DO UPDATE SET
                        finance_source = EXCLUDED.finance_source,
                        school_laestab = EXCLUDED.school_laestab,
                        school_name = EXCLUDED.school_name,
                        trust_uid = EXCLUDED.trust_uid,
                        trust_name = EXCLUDED.trust_name,
                        phase = EXCLUDED.phase,
                        overall_phase = EXCLUDED.overall_phase,
                        admissions_policy = EXCLUDED.admissions_policy,
                        urban_rural = EXCLUDED.urban_rural,
                        pupils_fte = EXCLUDED.pupils_fte,
                        teachers_fte = EXCLUDED.teachers_fte,
                        fsm_pct = EXCLUDED.fsm_pct,
                        ehcp_pct = EXCLUDED.ehcp_pct,
                        sen_support_pct = EXCLUDED.sen_support_pct,
                        eal_pct = EXCLUDED.eal_pct,
                        total_grant_funding_gbp = EXCLUDED.total_grant_funding_gbp,
                        total_self_generated_funding_gbp = EXCLUDED.total_self_generated_funding_gbp,
                        total_income_gbp = EXCLUDED.total_income_gbp,
                        teaching_staff_costs_gbp = EXCLUDED.teaching_staff_costs_gbp,
                        supply_teaching_staff_costs_gbp = EXCLUDED.supply_teaching_staff_costs_gbp,
                        education_support_staff_costs_gbp = EXCLUDED.education_support_staff_costs_gbp,
                        other_staff_costs_gbp = EXCLUDED.other_staff_costs_gbp,
                        total_staff_costs_gbp = EXCLUDED.total_staff_costs_gbp,
                        maintenance_improvement_costs_gbp = EXCLUDED.maintenance_improvement_costs_gbp,
                        premises_costs_gbp = EXCLUDED.premises_costs_gbp,
                        educational_supplies_costs_gbp = EXCLUDED.educational_supplies_costs_gbp,
                        bought_in_professional_services_costs_gbp =
                            EXCLUDED.bought_in_professional_services_costs_gbp,
                        catering_costs_gbp = EXCLUDED.catering_costs_gbp,
                        total_expenditure_gbp = EXCLUDED.total_expenditure_gbp,
                        revenue_reserve_gbp = EXCLUDED.revenue_reserve_gbp,
                        in_year_balance_gbp = EXCLUDED.in_year_balance_gbp,
                        source_file_url = EXCLUDED.source_file_url,
                        source_updated_at_utc = EXCLUDED.source_updated_at_utc
                    """
                )
                for staged_chunk in chunked(staged_rows, context.stage_chunk_size):
                    connection.execute(insert, [row.__dict__ for row in staged_chunk])

            if rejected_rows:
                rejection_insert = text(
                    """
                    INSERT INTO pipeline_rejections (
                        run_id,
                        source,
                        stage,
                        reason_code,
                        raw_record
                    ) VALUES (
                        :run_id,
                        :source,
                        'stage',
                        :reason_code,
                        CAST(:raw_record AS jsonb)
                    )
                    """
                )
                for rejected_chunk in chunked(rejected_rows, context.stage_chunk_size):
                    connection.execute(
                        rejection_insert,
                        [
                            {
                                "run_id": str(context.run_id),
                                "source": context.source.value,
                                "reason_code": reason_code,
                                "raw_record": json.dumps(raw_row, ensure_ascii=True, default=str),
                            }
                            for reason_code, raw_row in rejected_chunk
                        ],
                    )

        return StageResult(
            staged_rows=len(staged_rows),
            rejected_rows=len(rejected_rows),
            contract_version=finance_contract.CONTRACT_VERSION,
        )

    def promote(self, context: PipelineRunContext) -> int:
        if self._engine is None:
            raise ValueError("Pipeline engine is required for promote.")

        staging_table_name = self._staging_table_name(context)
        with self._engine.begin() as connection:
            promoted_rows = int(
                connection.execute(
                    text(
                        f"""
                        WITH upserted AS (
                            INSERT INTO school_financials_yearly (
                                urn,
                                academic_year,
                                finance_source,
                                school_laestab,
                                school_name,
                                trust_uid,
                                trust_name,
                                phase,
                                overall_phase,
                                admissions_policy,
                                urban_rural,
                                pupils_fte,
                                teachers_fte,
                                fsm_pct,
                                ehcp_pct,
                                sen_support_pct,
                                eal_pct,
                                total_grant_funding_gbp,
                                total_self_generated_funding_gbp,
                                total_income_gbp,
                                teaching_staff_costs_gbp,
                                supply_teaching_staff_costs_gbp,
                                education_support_staff_costs_gbp,
                                other_staff_costs_gbp,
                                total_staff_costs_gbp,
                                maintenance_improvement_costs_gbp,
                                premises_costs_gbp,
                                educational_supplies_costs_gbp,
                                bought_in_professional_services_costs_gbp,
                                catering_costs_gbp,
                                total_expenditure_gbp,
                                revenue_reserve_gbp,
                                in_year_balance_gbp,
                                income_per_pupil_gbp,
                                expenditure_per_pupil_gbp,
                                staff_costs_pct_of_expenditure,
                                teaching_staff_costs_per_pupil_gbp,
                                supply_staff_costs_pct_of_staff_costs,
                                revenue_reserve_per_pupil_gbp,
                                source_file_url,
                                source_updated_at_utc,
                                updated_at
                            )
                            SELECT
                                staged.urn,
                                staged.academic_year,
                                staged.finance_source,
                                staged.school_laestab,
                                staged.school_name,
                                staged.trust_uid,
                                staged.trust_name,
                                staged.phase,
                                staged.overall_phase,
                                staged.admissions_policy,
                                staged.urban_rural,
                                staged.pupils_fte,
                                staged.teachers_fte,
                                staged.fsm_pct,
                                staged.ehcp_pct,
                                staged.sen_support_pct,
                                staged.eal_pct,
                                staged.total_grant_funding_gbp,
                                staged.total_self_generated_funding_gbp,
                                staged.total_income_gbp,
                                staged.teaching_staff_costs_gbp,
                                staged.supply_teaching_staff_costs_gbp,
                                staged.education_support_staff_costs_gbp,
                                staged.other_staff_costs_gbp,
                                staged.total_staff_costs_gbp,
                                staged.maintenance_improvement_costs_gbp,
                                staged.premises_costs_gbp,
                                staged.educational_supplies_costs_gbp,
                                staged.bought_in_professional_services_costs_gbp,
                                staged.catering_costs_gbp,
                                staged.total_expenditure_gbp,
                                staged.revenue_reserve_gbp,
                                staged.in_year_balance_gbp,
                                CASE
                                    WHEN staged.pupils_fte IS NULL OR staged.pupils_fte <= 0
                                         OR staged.total_income_gbp IS NULL
                                        THEN NULL
                                    ELSE staged.total_income_gbp / staged.pupils_fte
                                END AS income_per_pupil_gbp,
                                CASE
                                    WHEN staged.pupils_fte IS NULL OR staged.pupils_fte <= 0
                                         OR staged.total_expenditure_gbp IS NULL
                                        THEN NULL
                                    ELSE staged.total_expenditure_gbp / staged.pupils_fte
                                END AS expenditure_per_pupil_gbp,
                                CASE
                                    WHEN staged.total_expenditure_gbp IS NULL
                                         OR staged.total_expenditure_gbp <= 0
                                         OR staged.total_staff_costs_gbp IS NULL
                                        THEN NULL
                                    ELSE staged.total_staff_costs_gbp / staged.total_expenditure_gbp
                                END AS staff_costs_pct_of_expenditure,
                                CASE
                                    WHEN staged.pupils_fte IS NULL OR staged.pupils_fte <= 0
                                         OR staged.teaching_staff_costs_gbp IS NULL
                                        THEN NULL
                                    ELSE staged.teaching_staff_costs_gbp / staged.pupils_fte
                                END AS teaching_staff_costs_per_pupil_gbp,
                                CASE
                                    WHEN staged.total_staff_costs_gbp IS NULL
                                         OR staged.total_staff_costs_gbp <= 0
                                         OR staged.supply_teaching_staff_costs_gbp IS NULL
                                        THEN NULL
                                    ELSE staged.supply_teaching_staff_costs_gbp / staged.total_staff_costs_gbp
                                END AS supply_staff_costs_pct_of_staff_costs,
                                CASE
                                    WHEN staged.pupils_fte IS NULL OR staged.pupils_fte <= 0
                                         OR staged.revenue_reserve_gbp IS NULL
                                        THEN NULL
                                    ELSE staged.revenue_reserve_gbp / staged.pupils_fte
                                END AS revenue_reserve_per_pupil_gbp,
                                staged.source_file_url,
                                staged.source_updated_at_utc,
                                timezone('utc', now())
                            FROM staging.{staging_table_name} AS staged
                            INNER JOIN schools
                                ON schools.urn = staged.urn
                            ON CONFLICT (urn, academic_year) DO UPDATE SET
                                finance_source = EXCLUDED.finance_source,
                                school_laestab = EXCLUDED.school_laestab,
                                school_name = EXCLUDED.school_name,
                                trust_uid = EXCLUDED.trust_uid,
                                trust_name = EXCLUDED.trust_name,
                                phase = EXCLUDED.phase,
                                overall_phase = EXCLUDED.overall_phase,
                                admissions_policy = EXCLUDED.admissions_policy,
                                urban_rural = EXCLUDED.urban_rural,
                                pupils_fte = EXCLUDED.pupils_fte,
                                teachers_fte = EXCLUDED.teachers_fte,
                                fsm_pct = EXCLUDED.fsm_pct,
                                ehcp_pct = EXCLUDED.ehcp_pct,
                                sen_support_pct = EXCLUDED.sen_support_pct,
                                eal_pct = EXCLUDED.eal_pct,
                                total_grant_funding_gbp = EXCLUDED.total_grant_funding_gbp,
                                total_self_generated_funding_gbp =
                                    EXCLUDED.total_self_generated_funding_gbp,
                                total_income_gbp = EXCLUDED.total_income_gbp,
                                teaching_staff_costs_gbp = EXCLUDED.teaching_staff_costs_gbp,
                                supply_teaching_staff_costs_gbp =
                                    EXCLUDED.supply_teaching_staff_costs_gbp,
                                education_support_staff_costs_gbp =
                                    EXCLUDED.education_support_staff_costs_gbp,
                                other_staff_costs_gbp = EXCLUDED.other_staff_costs_gbp,
                                total_staff_costs_gbp = EXCLUDED.total_staff_costs_gbp,
                                maintenance_improvement_costs_gbp =
                                    EXCLUDED.maintenance_improvement_costs_gbp,
                                premises_costs_gbp = EXCLUDED.premises_costs_gbp,
                                educational_supplies_costs_gbp =
                                    EXCLUDED.educational_supplies_costs_gbp,
                                bought_in_professional_services_costs_gbp =
                                    EXCLUDED.bought_in_professional_services_costs_gbp,
                                catering_costs_gbp = EXCLUDED.catering_costs_gbp,
                                total_expenditure_gbp = EXCLUDED.total_expenditure_gbp,
                                revenue_reserve_gbp = EXCLUDED.revenue_reserve_gbp,
                                in_year_balance_gbp = EXCLUDED.in_year_balance_gbp,
                                income_per_pupil_gbp = EXCLUDED.income_per_pupil_gbp,
                                expenditure_per_pupil_gbp = EXCLUDED.expenditure_per_pupil_gbp,
                                staff_costs_pct_of_expenditure =
                                    EXCLUDED.staff_costs_pct_of_expenditure,
                                teaching_staff_costs_per_pupil_gbp =
                                    EXCLUDED.teaching_staff_costs_per_pupil_gbp,
                                supply_staff_costs_pct_of_staff_costs =
                                    EXCLUDED.supply_staff_costs_pct_of_staff_costs,
                                revenue_reserve_per_pupil_gbp =
                                    EXCLUDED.revenue_reserve_per_pupil_gbp,
                                source_file_url = EXCLUDED.source_file_url,
                                source_updated_at_utc = EXCLUDED.source_updated_at_utc,
                                updated_at = timezone('utc', now())
                            RETURNING 1
                        )
                        SELECT COUNT(*) FROM upserted
                        """
                    )
                ).scalar_one()
            )
            connection.execute(text(f"DROP TABLE IF EXISTS staging.{staging_table_name}"))
        return promoted_rows

    @staticmethod
    def _staging_table_name(context: PipelineRunContext) -> str:
        return f"school_financial_benchmarks__{context.run_id.hex}"


def _normalize_source_url(value: str) -> str:
    normalized = value.strip()
    if not normalized:
        raise ValueError("Workbook URL must not be empty.")
    return normalized


def _build_bronze_file_name(source_url: str) -> str:
    parsed = urlsplit(source_url)
    candidate = Path(parsed.path).name if parsed.scheme else Path(source_url).name
    normalized = candidate.strip().replace(" ", "_")
    return normalized or "school_financial_benchmarks.xlsx"


def _clear_existing_bronze_files(bronze_source_path: Path) -> None:
    for path in bronze_source_path.glob("*"):
        if path.is_file():
            path.unlink()


def _required_str(value: object, message: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(message)
    return value.strip()


def _parse_row_count(value: object) -> int:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    raise ValueError("row_count is missing or invalid")


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _sha256_file(path: Path) -> str:
    return _sha256_bytes(path.read_bytes())


def _parse_datetime(value: str) -> datetime:
    return datetime.fromisoformat(value.replace("Z", "+00:00"))


def _academic_year_sort_key(value: str) -> tuple[int, str]:
    try:
        start_year = int(value.split("/", maxsplit=1)[0])
    except (ValueError, IndexError):
        start_year = -1
    return start_year, value


def _parse_academic_year(source_url: str) -> str:
    candidate = source_url.strip()
    match = _ACADEMIC_YEAR_PATTERN.search(candidate)
    if match is None:
        raise ValueError(f"Unable to determine academic year from workbook URL '{source_url}'.")
    start_year = match.group("start")
    end_year = match.group("end")
    if len(end_year) == 4:
        end_year = end_year[-2:]
    return f"{start_year}/{end_year}"


def _download_workbook(source_url: str) -> bytes:
    parsed = urlsplit(source_url)
    if parsed.scheme in {"http", "https"}:
        request = urllib.request.Request(
            source_url,
            headers={"User-Agent": "civitas-pipeline/0.1"},
        )
        with urllib.request.urlopen(request, timeout=60) as response:
            return response.read()

    local_path = Path(source_url)
    if local_path.exists():
        return local_path.read_bytes()
    raise FileNotFoundError(f"Workbook source '{source_url}' was not found.")


def _worksheet_row_count(worksheet: Worksheet) -> int:
    return sum(1 for _ in _worksheet_rows(worksheet))


def _worksheet_rows(worksheet: Worksheet) -> Sequence[dict[str, object]]:
    iterator = worksheet.iter_rows(values_only=True)
    headers = _find_worksheet_headers(iterator)
    if headers is None:
        return ()

    rows: list[dict[str, object]] = []
    for values in iterator:
        row = {
            header: cast(object, value)
            for header, value in zip(headers, values, strict=False)
            if header
        }
        if len(row) == 0:
            continue
        rows.append(row)
    return rows


def _find_worksheet_headers(
    iterator: Iterable[tuple[object, ...]],
) -> list[str] | None:
    for values in iterator:
        headers = [str(value).strip() if value is not None else "" for value in values]
        if _REQUIRED_WORKSHEET_HEADERS.issubset(headers):
            return headers
    return None


def _load_manifest_assets(manifest_path: Path) -> tuple[_ManifestAsset, ...]:
    if not manifest_path.exists():
        return ()
    try:
        payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return ()
    assets_payload = payload.get("assets") if isinstance(payload, dict) else None
    if not isinstance(assets_payload, list):
        return ()

    assets: list[_ManifestAsset] = []
    for item in assets_payload:
        if not isinstance(item, dict):
            continue
        sheet_names = item.get("sheet_names")
        sheet_row_counts = item.get("sheet_row_counts")
        if not isinstance(sheet_names, list) or not isinstance(sheet_row_counts, dict):
            continue
        try:
            assets.append(
                _ManifestAsset(
                    source_url=_required_str(item.get("source_url"), "missing source_url"),
                    bronze_file_name=_required_str(
                        item.get("bronze_file_name"),
                        "missing bronze_file_name",
                    ),
                    downloaded_at=_required_str(item.get("downloaded_at"), "missing downloaded_at"),
                    sha256=_required_str(item.get("sha256"), "missing sha256"),
                    academic_year=_required_str(item.get("academic_year"), "missing academic_year"),
                    sheet_names=tuple(
                        _required_str(sheet_name, "missing sheet name")
                        for sheet_name in sheet_names
                    ),
                    sheet_row_counts={
                        str(sheet_name): _parse_row_count(row_count)
                        for sheet_name, row_count in sheet_row_counts.items()
                    },
                )
            )
        except (TypeError, ValueError):
            continue
    return tuple(sorted(assets, key=lambda item: (item.academic_year, item.bronze_file_name)))


def _manifest_assets_exist(
    assets: Sequence[_ManifestAsset],
    *,
    bronze_source_path: Path,
) -> bool:
    if len(assets) == 0:
        return False
    for asset in assets:
        path = bronze_source_path / asset.bronze_file_name
        if not path.exists():
            return False
        if _sha256_file(path) != asset.sha256:
            return False
    return True
