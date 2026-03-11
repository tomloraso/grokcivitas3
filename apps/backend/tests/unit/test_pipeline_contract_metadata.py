from __future__ import annotations

import io
import json
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import parse_qs, urlsplit
from uuid import uuid4

from openpyxl import Workbook

from civitas.infrastructure.pipelines.base import PipelineRunContext, PipelineSource
from civitas.infrastructure.pipelines.contracts import (
    dfe_attendance as attendance_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    dfe_behaviour as behaviour_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    dfe_performance as performance_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    dfe_workforce as workforce_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    gias as gias_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    ks4_subject_performance as ks4_subject_performance_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    leaver_destinations as leaver_destinations_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    ofsted_latest as ofsted_latest_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    ofsted_timeline as ofsted_timeline_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    ons_imd as ons_imd_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    police as police_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    school_admissions as admissions_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    school_financial_benchmarks as finance_contract,
)
from civitas.infrastructure.pipelines.contracts import (
    sixteen_to_eighteen_subject_performance as sixteen_to_eighteen_subject_performance_contract,
)
from civitas.infrastructure.pipelines.demographics_release_files import (
    BRONZE_MANIFEST_FILE_NAME as DEMOGRAPHICS_MANIFEST_FILE_NAME,
)
from civitas.infrastructure.pipelines.demographics_release_files import (
    DemographicsReleaseFilesPipeline,
)
from civitas.infrastructure.pipelines.dfe_attendance import (
    BRONZE_MANIFEST_FILE_NAME as DFE_ATTENDANCE_MANIFEST_FILE_NAME,
)
from civitas.infrastructure.pipelines.dfe_attendance import (
    DfeAttendancePipeline,
)
from civitas.infrastructure.pipelines.dfe_behaviour import (
    BRONZE_MANIFEST_FILE_NAME as DFE_BEHAVIOUR_MANIFEST_FILE_NAME,
)
from civitas.infrastructure.pipelines.dfe_behaviour import (
    DfeBehaviourPipeline,
)
from civitas.infrastructure.pipelines.dfe_performance import (
    BRONZE_MANIFEST_FILE_NAME as DFE_PERFORMANCE_MANIFEST_FILE_NAME,
)
from civitas.infrastructure.pipelines.dfe_performance import (
    DfePerformancePipeline,
)
from civitas.infrastructure.pipelines.dfe_workforce import (
    BRONZE_MANIFEST_FILE_NAME as DFE_WORKFORCE_MANIFEST_FILE_NAME,
)
from civitas.infrastructure.pipelines.dfe_workforce import (
    DfeWorkforcePipeline,
)
from civitas.infrastructure.pipelines.gias import (
    BRONZE_FILE_NAME as GIAS_BRONZE_FILE_NAME,
)
from civitas.infrastructure.pipelines.gias import (
    GiasPipeline,
)
from civitas.infrastructure.pipelines.ks4_subject_performance import (
    BRONZE_MANIFEST_FILE_NAME as KS4_SUBJECT_PERFORMANCE_MANIFEST_FILE_NAME,
)
from civitas.infrastructure.pipelines.ks4_subject_performance import (
    Ks4SubjectPerformancePipeline,
)
from civitas.infrastructure.pipelines.leaver_destinations import (
    BRONZE_MANIFEST_FILE_NAME as LEAVER_DESTINATIONS_MANIFEST_FILE_NAME,
)
from civitas.infrastructure.pipelines.leaver_destinations import (
    LeaverDestinationsPipeline,
)
from civitas.infrastructure.pipelines.ofsted_latest import (
    BRONZE_FILE_NAME as OFSTED_LATEST_BRONZE_FILE_NAME,
)
from civitas.infrastructure.pipelines.ofsted_latest import (
    OfstedLatestPipeline,
)
from civitas.infrastructure.pipelines.ofsted_timeline import (
    BRONZE_MANIFEST_FILE_NAME,
    OfstedTimelinePipeline,
)
from civitas.infrastructure.pipelines.ons_imd import (
    BRONZE_METADATA_FILE_NAME as ONS_METADATA_FILE_NAME,
)
from civitas.infrastructure.pipelines.ons_imd import (
    OnsImdPipeline,
)
from civitas.infrastructure.pipelines.police_crime_context import (
    BRONZE_METADATA_FILE_NAME as POLICE_METADATA_FILE_NAME,
)
from civitas.infrastructure.pipelines.police_crime_context import (
    PoliceCrimeContextPipeline,
)
from civitas.infrastructure.pipelines.school_admissions import (
    BRONZE_MANIFEST_FILE_NAME as SCHOOL_ADMISSIONS_MANIFEST_FILE_NAME,
)
from civitas.infrastructure.pipelines.school_admissions import (
    SchoolAdmissionsPipeline,
)
from civitas.infrastructure.pipelines.school_financial_benchmarks import (
    BRONZE_MANIFEST_FILE_NAME as SCHOOL_FINANCIAL_BENCHMARKS_MANIFEST_FILE_NAME,
)
from civitas.infrastructure.pipelines.school_financial_benchmarks import (
    SchoolFinancialBenchmarksPipeline,
)
from civitas.infrastructure.pipelines.sixteen_to_eighteen_subject_performance import (
    BRONZE_MANIFEST_FILE_NAME as SIXTEEN_TO_EIGHTEEN_SUBJECT_PERFORMANCE_MANIFEST_FILE_NAME,
)
from civitas.infrastructure.pipelines.sixteen_to_eighteen_subject_performance import (
    SixteenToEighteenSubjectPerformancePipeline,
)

FIXTURES_ROOT = Path(__file__).resolve().parent.parent / "fixtures"


def _context(source: PipelineSource, bronze_root: Path) -> PipelineRunContext:
    return PipelineRunContext(
        run_id=uuid4(),
        source=source,
        started_at=datetime(2026, 3, 3, 10, 0, tzinfo=timezone.utc),
        bronze_root=bronze_root,
    )


def test_gias_download_writes_contract_version_to_metadata(tmp_path: Path) -> None:
    fixture = FIXTURES_ROOT / "gias" / "edubasealldata_valid.csv"
    pipeline = GiasPipeline(engine=None, source_csv=str(fixture))
    context = _context(PipelineSource.GIAS, tmp_path / "bronze")

    pipeline.download(context)

    metadata_path = context.bronze_source_path / GIAS_BRONZE_FILE_NAME
    payload = json.loads(metadata_path.with_suffix(".metadata.json").read_text(encoding="utf-8"))
    assert payload["normalization_contract_version"] == gias_contract.CONTRACT_VERSION


def test_demographics_download_writes_contract_version_to_manifest(tmp_path: Path) -> None:
    spc_release_html = (
        '<html><script id="__NEXT_DATA__" type="application/json">'
        '{"props":{"pageProps":{"releaseVersion":{"id":"spc-rv-1","downloadFiles":[{"id":"spc-file-1","name":"School level underlying data 2025"}]}}}}'
        "</script></html>"
    )
    sen_release_html = (
        '<html><script id="__NEXT_DATA__" type="application/json">'
        '{"props":{"pageProps":{"releaseVersion":{"id":"sen-rv-1","downloadFiles":[{"id":"sen-file-1","name":"School level underlying data 2025"}]}}}}'
        "</script></html>"
    )
    spc_csv = (
        "urn,time_period,% of pupils known to be eligible for free school meals (Performance Tables),"
        "% of pupils whose first language is known or believed to be other than English,"
        "% of pupils whose first language is known or believed to be English,"
        "% of pupils whose first language is unclassified\n"
        "100001,202425,18.2,8.4,90.6,1.0\n"
    )
    sen_csv = "URN,time_period,Total pupils,SEN support,EHC plan\n100001,202425,240,36,9\n"
    responses = {
        "https://explore-education-statistics.service.gov.uk/find-statistics/school-pupils-and-their-characteristics/2024-25": spc_release_html,
        "https://explore-education-statistics.service.gov.uk/find-statistics/special-educational-needs-in-england/2024-25": sen_release_html,
        "https://content.explore-education-statistics.service.gov.uk/api/releases/spc-rv-1/files/spc-file-1": spc_csv,
        "https://content.explore-education-statistics.service.gov.uk/api/releases/sen-rv-1/files/sen-file-1": sen_csv,
    }

    def fetcher(url: str) -> str:
        if url not in responses:
            raise AssertionError(f"Unexpected URL: {url}")
        return responses[url]

    pipeline = DemographicsReleaseFilesPipeline(
        engine=None,
        spc_publication_slug="school-pupils-and-their-characteristics",
        sen_publication_slug="special-educational-needs-in-england",
        release_slugs=("2024-25",),
        lookback_years=1,
        fetcher=fetcher,
    )
    context = _context(PipelineSource.DFE_CHARACTERISTICS, tmp_path / "bronze")

    pipeline.download(context)

    manifest_path = context.bronze_source_path / DEMOGRAPHICS_MANIFEST_FILE_NAME
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert payload["normalization_contract_version"] == "demographics_release_files.v1"
    assert len(payload["assets"]) == 2


def test_dfe_performance_download_writes_contract_version_to_manifest(
    tmp_path: Path,
    monkeypatch,
) -> None:
    ks2_dataset_id = "ks2-dataset"
    ks4_dataset_id = "ks4-dataset"

    def fake_download_json(url: str, *, timeout_seconds: float) -> dict[str, object]:
        parsed_url = urlsplit(url)
        query = parse_qs(parsed_url.query)
        if parsed_url.path.endswith(f"/data-sets/{ks2_dataset_id}/meta"):
            assert query.get("dataSetVersion") == ["ks2-v1"]
            return {
                "locations": {
                    "options": [
                        {"id": "100001", "code": "100001"},
                    ]
                }
            }
        if parsed_url.path.endswith(f"/data-sets/{ks4_dataset_id}/meta"):
            assert query.get("dataSetVersion") == ["ks4-v1"]
            return {
                "locations": {
                    "options": [
                        {"id": "100001", "code": "100001"},
                    ]
                }
            }
        if parsed_url.path.endswith(f"/data-sets/{ks2_dataset_id}"):
            assert query == {}
            return {"latestVersion": {"version": "ks2-v1"}}
        if parsed_url.path.endswith(f"/data-sets/{ks4_dataset_id}"):
            assert query == {}
            return {"latestVersion": {"version": "ks4-v1"}}
        raise AssertionError(f"Unexpected URL: {url}")

    def fake_post_json(
        url: str,
        *,
        payload: dict[str, object],
        timeout_seconds: float,
    ) -> dict[str, object]:
        parsed_url = urlsplit(url)
        query = parse_qs(parsed_url.query)
        assert payload["page"] == 1
        assert payload["pageSize"] == 10_000
        if parsed_url.path.endswith(f"/data-sets/{ks2_dataset_id}/query"):
            assert query.get("dataSetVersion") == ["ks2-v1"]
            assert payload["indicators"] == list(performance_contract.KS2_INDICATOR_IDS.values())
            return {
                "results": [
                    {
                        "timePeriod": {"period": "2024/2025"},
                        "locations": {"SCH": "100001"},
                        "filters": {"fV8YF": "EXcPq", "jfhAM": "2id7l"},
                        "values": {"IwjBz": "74.0", "i2s6X": "12.0"},
                    }
                ],
                "paging": {"totalPages": 1},
            }
        if parsed_url.path.endswith(f"/data-sets/{ks4_dataset_id}/query"):
            assert query.get("dataSetVersion") == ["ks4-v1"]
            assert payload["indicators"] == list(performance_contract.KS4_INDICATOR_IDS.values())
            return {
                "results": [
                    {
                        "timePeriod": {"period": "2024/2025"},
                        "locations": {"SCH": "100001"},
                        "filters": {
                            "pPmSo": "5Kydi",
                            "IzpBz": "mws9K",
                            "ibG6X": "WCb2b",
                            "LZ6Wj": "9b64v",
                        },
                        "values": {
                            "kgVhs": "47.2",
                            "Pwoeb": "0.11",
                            "dDo0Z": "52.3",
                            "hCRyW": "71.4",
                            "bmztT": "36.2",
                            "uEko4": "25.5",
                            "mqo9K": "31.3",
                        },
                    }
                ],
                "paging": {"totalPages": 1},
            }
        raise AssertionError(f"Unexpected URL: {url}")

    monkeypatch.setattr(
        "civitas.infrastructure.pipelines.dfe_performance._download_json",
        fake_download_json,
    )
    monkeypatch.setattr(
        "civitas.infrastructure.pipelines.dfe_performance._post_json",
        fake_post_json,
    )

    pipeline = DfePerformancePipeline(
        engine=None,
        ks2_dataset_id=ks2_dataset_id,
        ks4_dataset_id=ks4_dataset_id,
    )
    context = _context(PipelineSource.DFE_PERFORMANCE, tmp_path / "bronze")

    downloaded_rows = pipeline.download(context)

    manifest_path = context.bronze_source_path / DFE_PERFORMANCE_MANIFEST_FILE_NAME
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert downloaded_rows == 2
    assert payload["normalization_contract_version"] == performance_contract.CONTRACT_VERSION
    assert len(payload["datasets"]) == 2
    assert len(payload["assets"]) == 4


def test_dfe_attendance_download_writes_contract_version_to_manifest(tmp_path: Path) -> None:
    release_html = (
        '<html><script id="__NEXT_DATA__" type="application/json">'
        '{"props":{"pageProps":{"releaseVersion":{"id":"attendance-rv-1","downloadFiles":[{"id":"attendance-file-1","name":"School level absence data"}]}}}}'
        "</script></html>"
    )
    csv_payload = (
        "school_urn,time_period,geographic_level,sess_overall_percent,enrolments_pa_10_exact_percent\n"
        "100001,202324,School,6.0,14.2\n"
    )
    responses = {
        "https://explore-education-statistics.service.gov.uk/find-statistics/pupil-absence-in-schools-in-england/2023-24": release_html,
        "https://content.explore-education-statistics.service.gov.uk/api/releases/attendance-rv-1/files/attendance-file-1": csv_payload,
    }

    def fetcher(url: str) -> str:
        if url not in responses:
            raise AssertionError(f"Unexpected URL: {url}")
        return responses[url]

    pipeline = DfeAttendancePipeline(
        engine=None,
        publication_slug="pupil-absence-in-schools-in-england",
        release_slugs=("2023-24",),
        lookback_years=1,
        fetcher=fetcher,
    )
    context = _context(PipelineSource.DFE_ATTENDANCE, tmp_path / "bronze")

    pipeline.download(context)

    manifest_path = context.bronze_source_path / DFE_ATTENDANCE_MANIFEST_FILE_NAME
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert payload["normalization_contract_version"] == attendance_contract.CONTRACT_VERSION
    assert len(payload["assets"]) == 1


def test_dfe_behaviour_download_writes_contract_version_to_manifest(tmp_path: Path) -> None:
    release_html = (
        '<html><script id="__NEXT_DATA__" type="application/json">'
        '{"props":{"pageProps":{"releaseVersion":{"id":"behaviour-rv-1","downloadFiles":[{"id":"behaviour-file-1","name":"School level suspensions and permanent exclusions data"}]}}}}'
        "</script></html>"
    )
    csv_payload = (
        "school_urn,time_period,geographic_level,suspension,susp_rate,perm_excl,perm_excl_rate\n"
        "100001,202324,School,121,16.4,1,0.1\n"
    )
    responses = {
        "https://explore-education-statistics.service.gov.uk/find-statistics/suspensions-and-permanent-exclusions-in-england/2024-25-autumn-term": release_html,
        "https://content.explore-education-statistics.service.gov.uk/api/releases/behaviour-rv-1/files/behaviour-file-1": csv_payload,
    }

    def fetcher(url: str) -> str:
        if url not in responses:
            raise AssertionError(f"Unexpected URL: {url}")
        return responses[url]

    pipeline = DfeBehaviourPipeline(
        engine=None,
        publication_slug="suspensions-and-permanent-exclusions-in-england",
        release_slugs=("2024-25-autumn-term",),
        lookback_years=1,
        fetcher=fetcher,
    )
    context = _context(PipelineSource.DFE_BEHAVIOUR, tmp_path / "bronze")

    pipeline.download(context)

    manifest_path = context.bronze_source_path / DFE_BEHAVIOUR_MANIFEST_FILE_NAME
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert payload["normalization_contract_version"] == behaviour_contract.CONTRACT_VERSION
    assert len(payload["assets"]) == 1


def test_dfe_workforce_download_writes_contract_version_to_manifest(tmp_path: Path) -> None:
    def build_zip_payload(entries: dict[str, str]) -> bytes:
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w") as archive:
            for name, payload in entries.items():
                archive.writestr(name, payload)
        return buffer.getvalue()

    release_html = (
        '<html><script id="__NEXT_DATA__" type="application/json">'
        '{"props":{"pageProps":{"releaseVersion":{"id":"workforce-rv-1","downloadFiles":[{"id":"workforce-file-1","name":"School level workforce data"}]}}}}'
        "</script></html>"
    )
    legacy_csv_payload = (
        "school_urn,time_period,geographic_level,pupil_teacher_ratio,supply_teacher_pct\n"
        "100001,202324,School,16.3,2.4\n"
    )
    teacher_characteristics_zip = build_zip_payload(
        {
            "teacher_characteristics.csv": (
                "school_urn,time_period,characteristic_group,characteristic\n"
                "100001,202324,sex,Male\n"
            )
        }
    )
    support_staff_zip = build_zip_payload(
        {
            "support_staff.csv": (
                "school_urn,time_period,post,sex,ethnicity_major\n"
                "100001,202324,Teaching assistant,Female,White\n"
            )
        }
    )
    teacher_pay_csv = (
        "school_urn,time_period,teacher_headcount_all,teacher_average_mean_salary_gbp\n"
        "100001,202324,12,42350\n"
    )
    teacher_absence_csv = (
        "school_urn,time_period,teacher_absence_pct,teacher_absence_days_total\n"
        "100001,202324,3.2,27\n"
    )
    teacher_vacancies_csv = (
        "school_urn,time_period,teacher_vacancy_count,teacher_vacancy_rate\n100001,202324,1,4.1\n"
    )
    third_party_support_csv = (
        "school_urn,time_period,post,headcount\n100001,202324,Education welfare officer,2\n"
    )
    empty_csv_payload = "school_urn,time_period\n"
    source_catalog_release_version_id = "workforce-catalog-rv-1"
    responses = {
        "https://explore-education-statistics.service.gov.uk/find-statistics/school-workforce-in-england/2024": release_html,
        "https://content.explore-education-statistics.service.gov.uk/api/releases/workforce-rv-1/files/workforce-file-1": legacy_csv_payload,
        f"https://content.explore-education-statistics.service.gov.uk/api/releases/{source_catalog_release_version_id}/files/43ec3624-b83f-47e4-8941-2fd2fd6bfd3f": teacher_characteristics_zip,
        f"https://content.explore-education-statistics.service.gov.uk/api/releases/{source_catalog_release_version_id}/files/89cc4c08-611b-4dd6-a370-184a205fe9d6": support_staff_zip,
        f"https://content.explore-education-statistics.service.gov.uk/api/releases/{source_catalog_release_version_id}/files/05001215-c1c7-4210-f9db-08dd8e3a5799": teacher_pay_csv,
        f"https://content.explore-education-statistics.service.gov.uk/api/releases/{source_catalog_release_version_id}/files/9be449b9-57cf-4199-9410-08dd97c8935b": teacher_absence_csv,
        f"https://content.explore-education-statistics.service.gov.uk/api/releases/{source_catalog_release_version_id}/files/0edf6802-1d84-4a9d-f9bd-08dd8e3a5799": teacher_vacancies_csv,
        f"https://content.explore-education-statistics.service.gov.uk/api/releases/{source_catalog_release_version_id}/files/d9fbbe2a-8106-452f-f9a7-08dd8e3a5799": third_party_support_csv,
        f"https://content.explore-education-statistics.service.gov.uk/api/releases/{source_catalog_release_version_id}/files/ed1c5650-9e67-453d-0d91-08ddcde6ffdc": empty_csv_payload,
        f"https://content.explore-education-statistics.service.gov.uk/api/releases/{source_catalog_release_version_id}/files/8aab402f-9fc3-44bf-c16d-08ddd67d4d79": empty_csv_payload,
    }

    def fetcher(url: str) -> str:
        if url not in responses:
            raise AssertionError(f"Unexpected URL: {url}")
        return responses[url]

    pipeline = DfeWorkforcePipeline(
        engine=None,
        publication_slug="school-workforce-in-england",
        release_slugs=("2024",),
        lookback_years=1,
        fetcher=fetcher,
        source_catalog_release_slug="2024",
        source_catalog_release_version_id=source_catalog_release_version_id,
    )
    context = _context(PipelineSource.DFE_WORKFORCE, tmp_path / "bronze")

    pipeline.download(context)

    manifest_path = context.bronze_source_path / DFE_WORKFORCE_MANIFEST_FILE_NAME
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert payload["normalization_contract_version"] == workforce_contract.CONTRACT_VERSION
    assert len(payload["assets"]) == 9
    assert {asset["asset_kind"] for asset in payload["assets"]} == {
        "legacy_workforce",
        "support_staff_characteristics",
        "teacher_absence",
        "teacher_characteristics",
        "teacher_pay",
        "teacher_turnover",
        "teacher_vacancies",
        "third_party_support",
        "workforce_size",
    }
    asset_statuses = {asset["asset_kind"]: asset["source_status"] for asset in payload["assets"]}
    assert asset_statuses["teacher_turnover"] == "empty"
    assert asset_statuses["workforce_size"] == "empty"


def test_school_admissions_download_writes_contract_version_to_manifest(tmp_path: Path) -> None:
    csv_path = tmp_path / "school_admissions.csv"
    csv_path.write_text(
        "time_period,time_identifier,geographic_level,country_code,country_name,region_code,"
        "region_name,old_la_code,new_la_code,la_name,school_phase,school_laestab_as_used,"
        "number_preferences_la,school_name,total_number_places_offered,"
        "number_preferred_offers,number_1st_preference_offers,number_2nd_preference_offers,"
        "number_3rd_preference_offers,times_put_as_any_preferred_school,"
        "times_put_as_1st_preference,times_put_as_2nd_preference,times_put_as_3rd_preference,"
        "proportion_1stprefs_v_1stprefoffers,proportion_1stprefs_v_totaloffers,"
        "all_applications_from_another_LA,offers_to_applicants_from_another_LA,"
        "establishment_type,denomination,FSM_eligible_percent,admissions_policy,urban_rural,"
        "allthrough_school,parliamentary_constituency_code,parliamentary_constituency_name,"
        "school_urn,entry_year\n"
        "202526,Academic year,School,E92000001,England,E12000007,London,213,E09000033,"
        "Westminster,Primary,2136007,6,Alpha Primary School,60,57,49,6,2,95,72,15,8,"
        "1.4694,1.2000,33,18,Community school,None,18.5,Comprehensive,"
        "Urban major conurbation,No,E14000639,Westminster,100001,R\n",
        encoding="utf-8",
    )

    pipeline = SchoolAdmissionsPipeline(engine=None, source_csv=str(csv_path))
    context = _context(PipelineSource.SCHOOL_ADMISSIONS, tmp_path / "bronze")

    pipeline.download(context)

    manifest_path = context.bronze_source_path / SCHOOL_ADMISSIONS_MANIFEST_FILE_NAME
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert payload["normalization_contract_version"] == admissions_contract.CONTRACT_VERSION
    assert len(payload["assets"]) == 1
    assert payload["assets"][0]["bronze_file_name"] == "school_admissions.csv"


def test_leaver_destinations_download_writes_contract_version_to_manifest(
    tmp_path: Path,
) -> None:
    ks4_csv_path = tmp_path / "ees_ks4_inst_202223.csv"
    ks4_csv_path.write_text(
        "time_period,time_identifier,geographic_level,country_code,country_name,region_code,"
        "region_name,old_la_code,new_la_code,la_name,local_authority_selection_status,"
        "school_laestab,school_urn,school_name,admission_policy,entry_gender,"
        "institution_group,institution_type,breakdown_topic,breakdown,data_type,version,"
        "cohort,overall,education,fe,ssf,sfc,other_edu,appren,all_work,all_notsust,"
        "all_unknown\n"
        "202223,Academic year,School,E92000001,England,E12000007,London,213,E09000033,"
        "Westminster,Selected,2136007,100001,Alpha School,Comprehensive,Mixed,"
        "Local authority maintained schools,Community school,Total,Total,Percentage,1,"
        "118,92.4,61.0,18.0,c,,4.0,7.3,17.5,3.1,4.5\n",
        encoding="utf-8",
    )
    study_csv_path = tmp_path / "ees_ks5_inst_202223.csv"
    study_csv_path.write_text(
        "time_period,time_identifier,geographic_level,country_code,country_name,region_code,"
        "region_name,old_la_code,new_la_code,la_name,local_authority_selection_status,"
        "school_laestab,school_urn,school_name,admission_policy,entry_gender,"
        "institution_group,institution_type,cohort_level_group,cohort_level,"
        "breakdown_topic,breakdown,data_type,version,cohort,overall,education,he,fe,"
        "other_edu,appren,all_work,all_notsust,all_unknown\n"
        "202223,Academic year,School,E92000001,England,E12000007,London,201,E09000001,"
        "City of London,Selected,2013614,100002,Beta Sixth Form,Not applicable,Mixed,"
        "Academies,Academy converter,Total,Total,Total,Total,Number of students,1,"
        "90,80,56,31,18,7,10,14,6,10\n",
        encoding="utf-8",
    )

    pipeline = LeaverDestinationsPipeline(
        engine=None,
        ks4_source_csv=str(ks4_csv_path),
        ks4_source_url="https://example.com/ks4.csv",
        ks4_release_page_url="https://example.com/ks4-release",
        ks4_data_catalogue_url="https://example.com/ks4-catalogue",
        study_16_to_18_source_csv=str(study_csv_path),
        study_16_to_18_source_url="https://example.com/16-to-18.csv",
        study_16_to_18_release_page_url="https://example.com/16-to-18-release",
        study_16_to_18_data_catalogue_url="https://example.com/16-to-18-catalogue",
    )
    context = _context(PipelineSource.LEAVER_DESTINATIONS, tmp_path / "bronze")

    downloaded_rows = pipeline.download(context)

    manifest_path = context.bronze_source_path / LEAVER_DESTINATIONS_MANIFEST_FILE_NAME
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert downloaded_rows == 2
    assert (
        payload["normalization_contract_version"] == leaver_destinations_contract.CONTRACT_VERSION
    )
    assert len(payload["assets"]) == 2
    assert {asset["destination_stage"] for asset in payload["assets"]} == {"ks4", "16_to_18"}
    assert payload["assets"][0]["public_csv_route_url"].startswith("https://example.com/")


def test_ks4_subject_performance_download_writes_contract_version_to_manifest(
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "ks4_subject_performance.csv"
    csv_path.write_text(
        "time_period,time_identifier,geographic_level,country_code,country_name,"
        "school_laestab,school_urn,school_name,old_la_code,new_la_code,la_name,version,"
        "establishment_type_group,pupil_count,qualification_type,qualification_detailed,"
        "grade_structure,subject,discount_code,subject_discount_group,grade,number_achieving\n"
        "202425,Academic year,School,E92000001,England,2136007,100001,Alpha School,213,"
        "E09000033,Westminster,Revised,Local authority maintained schools,118,GCSE,"
        "GCSE (9-1) Full Course,9 / 8 / 7 / 6 / 5 / 4 / 3 / 2 / 1 / U / X,"
        "Mathematics,MA1,Mathematics,7,24\n",
        encoding="utf-8",
    )

    pipeline = Ks4SubjectPerformancePipeline(
        engine=None,
        source_csv=str(csv_path),
        source_url="https://example.com/ks4-subject-performance.csv",
        release_page_url="https://example.com/ks4-subject-performance-release",
        data_catalogue_url="https://example.com/ks4-subject-performance-catalogue",
    )
    context = _context(PipelineSource.KS4_SUBJECT_PERFORMANCE, tmp_path / "bronze")

    pipeline.download(context)

    manifest_path = context.bronze_source_path / KS4_SUBJECT_PERFORMANCE_MANIFEST_FILE_NAME
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert (
        payload["normalization_contract_version"]
        == ks4_subject_performance_contract.CONTRACT_VERSION
    )
    assert len(payload["assets"]) == 1
    assert payload["assets"][0]["academic_year"] == "2024/25"
    assert (
        payload["assets"][0]["public_csv_route_url"]
        == "https://example.com/ks4-subject-performance.csv"
    )


def test_sixteen_to_eighteen_subject_performance_download_writes_contract_version_to_manifest(
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "sixteen_to_eighteen_subject_performance.csv"
    csv_path.write_text(
        "time_period,time_identifier,geographic_level,country_code,country_name,version,"
        "old_la_code,new_la_code,la_name,school_name,school_urn,school_laestab,exam_cohort,"
        "qualification_detailed,qualification_level,a_level_equivelent_size,"
        "gcse_equivelent_size,grade_structure,subject,grade,entries_count\n"
        "202425,Academic year,School,E92000001,England,Revised,201,E09000001,"
        "City of London,Beta Sixth Form,100002,2013614,A level,GCE A level,3,1,4,"
        "*,A,B,C,D,E,Mathematics,A,12\n",
        encoding="utf-8",
    )

    pipeline = SixteenToEighteenSubjectPerformancePipeline(
        engine=None,
        source_csv=str(csv_path),
        source_url="https://example.com/16-to-18-subject-performance.csv",
        release_page_url="https://example.com/16-to-18-subject-performance-release",
        data_catalogue_url="https://example.com/16-to-18-subject-performance-catalogue",
    )
    context = _context(PipelineSource.SIXTEEN_TO_EIGHTEEN_SUBJECT_PERFORMANCE, tmp_path / "bronze")

    pipeline.download(context)

    manifest_path = (
        context.bronze_source_path / SIXTEEN_TO_EIGHTEEN_SUBJECT_PERFORMANCE_MANIFEST_FILE_NAME
    )
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert (
        payload["normalization_contract_version"]
        == sixteen_to_eighteen_subject_performance_contract.CONTRACT_VERSION
    )
    assert len(payload["assets"]) == 1
    assert payload["assets"][0]["academic_year"] == "2024/25"
    assert (
        payload["assets"][0]["public_csv_route_url"]
        == "https://example.com/16-to-18-subject-performance.csv"
    )


def test_school_financial_benchmarks_download_writes_contract_version_to_manifest(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "AAR_2023-24_download.xlsx"
    workbook = Workbook()
    index_sheet = workbook.active
    index_sheet.title = "Index"
    index_sheet.append(["Sheet", "Rows"])
    index_sheet.append(["Academies", 1])
    academies_sheet = workbook.create_sheet("Academies")
    academies_sheet.append(["URN", "School Name", "Total Income"])
    academies_sheet.append(["100001", "Example Academy", 2070000])
    workbook.create_sheet("Central Services")
    workbook.save(workbook_path)

    pipeline = SchoolFinancialBenchmarksPipeline(
        engine=None,
        workbook_urls=(str(workbook_path),),
    )
    context = _context(PipelineSource.SCHOOL_FINANCIAL_BENCHMARKS, tmp_path / "bronze")

    pipeline.download(context)

    manifest_path = context.bronze_source_path / SCHOOL_FINANCIAL_BENCHMARKS_MANIFEST_FILE_NAME
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert payload["normalization_contract_version"] == finance_contract.CONTRACT_VERSION
    assert len(payload["assets"]) == 1
    assert payload["assets"][0]["academic_year"] == "2023/24"
    assert payload["assets"][0]["sheet_names"] == ["Index", "Academies", "Central Services"]
    assert payload["assets"][0]["sheet_row_counts"]["Academies"] == 1


def test_ofsted_latest_download_writes_contract_version_to_metadata(tmp_path: Path) -> None:
    fixture = FIXTURES_ROOT / "ofsted_latest" / "latest_inspections_valid.csv"
    pipeline = OfstedLatestPipeline(engine=None, source_csv=str(fixture))
    context = _context(PipelineSource.OFSTED_LATEST, tmp_path / "bronze")

    pipeline.download(context)

    metadata_path = context.bronze_source_path / OFSTED_LATEST_BRONZE_FILE_NAME
    payload = json.loads(metadata_path.with_suffix(".metadata.json").read_text(encoding="utf-8"))
    assert payload["normalization_contract_version"] == ofsted_latest_contract.CONTRACT_VERSION


def test_ofsted_timeline_download_writes_contract_version_to_manifest(tmp_path: Path) -> None:
    fixtures_root = FIXTURES_ROOT / "ofsted_timeline"
    ytd_fixture = fixtures_root / "all_inspections_ytd_mixed.csv"
    historical_fixture = fixtures_root / "all_inspections_historical_2015_2019_mixed.csv"

    pipeline = OfstedTimelinePipeline(
        engine=None,
        source_assets_csv=f"{ytd_fixture},{historical_fixture}",
    )
    context = _context(PipelineSource.OFSTED_TIMELINE, tmp_path / "bronze")

    pipeline.download(context)

    manifest_path = context.bronze_source_path / BRONZE_MANIFEST_FILE_NAME
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert payload["normalization_contract_version"] == ofsted_timeline_contract.CONTRACT_VERSION


def test_ons_download_writes_contract_version_to_metadata(tmp_path: Path) -> None:
    fixture = FIXTURES_ROOT / "ons_imd" / "file_7_valid_2025.csv"
    pipeline = OnsImdPipeline(engine=None, source_csv=str(fixture))
    context = _context(PipelineSource.ONS_IMD, tmp_path / "bronze")

    pipeline.download(context)

    metadata_path = context.bronze_source_path / ONS_METADATA_FILE_NAME
    payload = json.loads(metadata_path.read_text(encoding="utf-8"))
    assert payload["normalization_contract_version"] == ons_imd_contract.CONTRACT_VERSION


def test_police_download_writes_contract_version_to_metadata(tmp_path: Path) -> None:
    csv_fixture = FIXTURES_ROOT / "police_crime_context" / "2026-01-example-street.csv"
    archive_path = tmp_path / "crime.zip"
    with zipfile.ZipFile(archive_path, "w") as archive:
        archive.writestr("2026-01-example-street.csv", csv_fixture.read_bytes())

    pipeline = PoliceCrimeContextPipeline(
        engine=None,
        source_archive_url=str(archive_path),
        source_mode="archive",
    )
    context = _context(PipelineSource.POLICE_CRIME_CONTEXT, tmp_path / "bronze")

    pipeline.download(context)

    metadata_path = context.bronze_source_path / POLICE_METADATA_FILE_NAME
    payload = json.loads(metadata_path.read_text(encoding="utf-8"))
    assert payload["normalization_contract_version"] == police_contract.CONTRACT_VERSION
    assert payload["archive_months"] == ["2026-01"]
    assert payload["archive_forces"] == ["example"]
    assert payload["archive_month_count"] == 1
    assert payload["archive_force_count"] == 1
