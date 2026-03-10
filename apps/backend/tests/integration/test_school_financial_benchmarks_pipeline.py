from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

from civitas.infrastructure.config.settings import AppSettings
from civitas.infrastructure.pipelines.base import PipelineRunContext, PipelineSource
from civitas.infrastructure.pipelines.school_financial_benchmarks import (
    BRONZE_MANIFEST_FILE_NAME,
    SchoolFinancialBenchmarksPipeline,
)


def _database_url() -> str:
    return AppSettings().database.url


def _build_engine(database_url: str) -> Engine:
    if database_url.startswith("postgresql"):
        return create_engine(database_url, future=True, connect_args={"connect_timeout": 2})
    return create_engine(database_url, future=True)


def _database_available(database_url: str) -> bool:
    engine = _build_engine(database_url)
    try:
        with engine.connect() as connection:
            connection.execute(text("SELECT 1"))
        return True
    except Exception:
        return False
    finally:
        engine.dispose()


DATABASE_URL = _database_url()
DATABASE_AVAILABLE = _database_available(DATABASE_URL)
pytestmark = pytest.mark.skipif(
    not DATABASE_AVAILABLE,
    reason="Postgres database unavailable for school financial benchmarks pipeline integration test.",
)


@pytest.fixture()
def engine() -> Engine:
    engine = _build_engine(DATABASE_URL)
    _ensure_schema(engine)
    try:
        yield engine
    finally:
        _cleanup(engine)
        engine.dispose()


def _ensure_schema(engine: Engine) -> None:
    with engine.begin() as connection:
        connection.execute(text("CREATE EXTENSION IF NOT EXISTS postgis"))
        connection.execute(text("CREATE SCHEMA IF NOT EXISTS staging"))
        connection.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS pipeline_runs (
                    run_id uuid PRIMARY KEY,
                    source text NOT NULL,
                    status text NOT NULL,
                    started_at timestamptz NOT NULL,
                    finished_at timestamptz NULL,
                    bronze_path text NOT NULL,
                    downloaded_rows integer NOT NULL DEFAULT 0,
                    staged_rows integer NOT NULL DEFAULT 0,
                    promoted_rows integer NOT NULL DEFAULT 0,
                    rejected_rows integer NOT NULL DEFAULT 0,
                    error_message text NULL
                )
                """
            )
        )
        connection.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS pipeline_rejections (
                    id bigserial PRIMARY KEY,
                    run_id uuid NOT NULL REFERENCES pipeline_runs(run_id) ON DELETE CASCADE,
                    source text NOT NULL,
                    stage text NOT NULL,
                    reason_code text NOT NULL,
                    raw_record jsonb NULL,
                    created_at timestamptz NOT NULL DEFAULT timezone('utc', now())
                )
                """
            )
        )
        connection.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS schools (
                    urn text PRIMARY KEY,
                    name text NOT NULL,
                    phase text NULL,
                    type text NULL,
                    status text NULL,
                    postcode text NULL,
                    easting double precision NOT NULL,
                    northing double precision NOT NULL,
                    location geography(Point, 4326) NOT NULL,
                    capacity integer NULL,
                    pupil_count integer NULL,
                    open_date date NULL,
                    close_date date NULL,
                    updated_at timestamptz NOT NULL DEFAULT timezone('utc', now())
                )
                """
            )
        )
        connection.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS school_financials_yearly (
                    urn text NOT NULL REFERENCES schools(urn) ON DELETE CASCADE,
                    academic_year text NOT NULL,
                    finance_source text NOT NULL DEFAULT 'aar',
                    school_laestab text NULL,
                    school_name text NOT NULL,
                    trust_uid text NULL,
                    trust_name text NULL,
                    phase text NULL,
                    overall_phase text NULL,
                    admissions_policy text NULL,
                    urban_rural text NULL,
                    pupils_fte double precision NULL,
                    teachers_fte double precision NULL,
                    fsm_pct double precision NULL,
                    ehcp_pct double precision NULL,
                    sen_support_pct double precision NULL,
                    eal_pct double precision NULL,
                    total_grant_funding_gbp double precision NULL,
                    total_self_generated_funding_gbp double precision NULL,
                    total_income_gbp double precision NULL,
                    teaching_staff_costs_gbp double precision NULL,
                    supply_teaching_staff_costs_gbp double precision NULL,
                    education_support_staff_costs_gbp double precision NULL,
                    other_staff_costs_gbp double precision NULL,
                    total_staff_costs_gbp double precision NULL,
                    maintenance_improvement_costs_gbp double precision NULL,
                    premises_costs_gbp double precision NULL,
                    educational_supplies_costs_gbp double precision NULL,
                    bought_in_professional_services_costs_gbp double precision NULL,
                    catering_costs_gbp double precision NULL,
                    total_expenditure_gbp double precision NULL,
                    revenue_reserve_gbp double precision NULL,
                    in_year_balance_gbp double precision NULL,
                    income_per_pupil_gbp double precision NULL,
                    expenditure_per_pupil_gbp double precision NULL,
                    staff_costs_pct_of_expenditure double precision NULL,
                    teaching_staff_costs_per_pupil_gbp double precision NULL,
                    supply_staff_costs_pct_of_staff_costs double precision NULL,
                    revenue_reserve_per_pupil_gbp double precision NULL,
                    source_file_url text NOT NULL,
                    source_updated_at_utc timestamptz NOT NULL,
                    updated_at timestamptz NOT NULL DEFAULT timezone('utc', now()),
                    PRIMARY KEY (urn, academic_year)
                )
                """
            )
        )


def _cleanup(engine: Engine) -> None:
    with engine.begin() as connection:
        connection.execute(
            text("DELETE FROM school_financials_yearly WHERE urn IN ('100001', '100002')")
        )
        connection.execute(text("DELETE FROM schools WHERE urn IN ('100001', '100002')"))


def _seed_schools(engine: Engine) -> None:
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                INSERT INTO schools (
                    urn, name, phase, type, status, postcode,
                    easting, northing, location, updated_at
                ) VALUES
                (
                    '100001', 'Alpha Academy', 'Primary', 'Academy converter', 'Open', 'SW1A 1AA',
                    529090, 179645,
                    ST_GeogFromText('SRID=4326;POINT(-0.141 51.501)'),
                    timezone('utc', now())
                ),
                (
                    '100002', 'Beta Academy', 'Secondary', 'Academy sponsor led', 'Open', 'SW1A 2AA',
                    529200, 179700,
                    ST_GeogFromText('SRID=4326;POINT(-0.140 51.502)'),
                    timezone('utc', now())
                )
                ON CONFLICT (urn) DO NOTHING
                """
            )
        )


def _insert_run_row(engine: Engine, context: PipelineRunContext) -> None:
    with engine.begin() as connection:
        connection.execute(
            text(
                """
                INSERT INTO pipeline_runs (
                    run_id,
                    source,
                    status,
                    started_at,
                    bronze_path,
                    downloaded_rows,
                    staged_rows,
                    promoted_rows,
                    rejected_rows
                ) VALUES (
                    :run_id,
                    :source,
                    'running',
                    :started_at,
                    :bronze_path,
                    0,
                    0,
                    0,
                    0
                )
                ON CONFLICT (run_id) DO NOTHING
                """
            ),
            {
                "run_id": str(context.run_id),
                "source": context.source.value,
                "started_at": context.started_at,
                "bronze_path": str(context.bronze_source_path),
            },
        )


def _context(bronze_root: Path) -> PipelineRunContext:
    return PipelineRunContext(
        run_id=uuid4(),
        source=PipelineSource.SCHOOL_FINANCIAL_BENCHMARKS,
        started_at=datetime(2026, 3, 10, 10, 0, tzinfo=timezone.utc),
        bronze_root=bronze_root,
    )


def _write_manifest_and_workbook(context: PipelineRunContext) -> None:
    context.bronze_source_path.mkdir(parents=True, exist_ok=True)
    workbook_name = "AAR_2023-24_download.xlsx"
    workbook_path = context.bronze_source_path / workbook_name
    workbook = Workbook()
    index_sheet = workbook.active
    index_sheet.title = "Index"
    index_sheet.append(["Sheet", "Rows"])
    index_sheet.append(["Academies", 4])

    academies_sheet = workbook.create_sheet("Academies")
    academies_sheet.append(
        [
            "LAEstab",
            "URN",
            "School Name",
            "UID",
            "Trust or Company Name",
            "Overall Phase",
            "Phase",
            "Admissions policy",
            "Urban/Rural",
            "Number of pupils in academy (FTE)",
            "Number of teachers in academy (FTE)",
            "% of pupils eligible for FSM",
            "% of pupils with EHCP",
            "% of pupil with SEN support",
            "% of pupils with English as an additional language",
            "Total Grant Funding",
            "Total Self Generated Funding",
            "Total Income",
            "Teaching staff",
            "Supply teaching staff",
            "Education support staff",
            "Other staff",
            "Total Staff Costs",
            "Maintenance & Improvement Costs",
            "Premises Costs",
            "Total Costs of Educational Supplies",
            "Costs of Brought in Professional Services",
            "Catering Expenses",
            "Total Expenditure",
            "Revenue Reserve",
            "In year balance",
        ]
    )
    academies_sheet.append(
        [
            "213/6007",
            "100001",
            "Alpha Academy",
            "5712",
            "Example Trust",
            "Primary",
            "Primary",
            "Not applicable",
            "Urban major conurbation",
            400,
            24,
            15.0,
            2.0,
            11.0,
            9.0,
            2250000,
            150000,
            2400000,
            1150000,
            30000,
            330000,
            190000,
            1700000,
            95000,
            120000,
            55000,
            45000,
            70000,
            2280000,
            320000,
            120000,
        ]
    )
    academies_sheet.append(
        [
            "213/6008",
            "100002",
            "Beta Academy",
            "5713",
            "Example Trust",
            "Secondary",
            "Secondary",
            "Comprehensive",
            "Urban city and town",
            "",
            "SUPP",
            ".",
            "n/a",
            "x",
            "na",
            "",
            "SUPP",
            "n/a",
            ".",
            "x",
            "na",
            "SUPP",
            "",
            "n/a",
            "SUPP",
            "",
            ".",
            "x",
            "na",
            "",
            "SUPP",
        ]
    )
    academies_sheet.append(
        [
            "213/6009",
            "ABC123",
            "Invalid Academy",
            "5714",
            "Example Trust",
            "Primary",
            "Primary",
            "Not applicable",
            "Urban major conurbation",
            300,
            18,
            12.0,
            1.5,
            10.0,
            7.0,
            1500000,
            100000,
            1600000,
            780000,
            22000,
            210000,
            120000,
            1132000,
            50000,
            75000,
            42000,
            32000,
            40000,
            1500000,
            100000,
            50000,
        ]
    )
    academies_sheet.append(
        [
            "213/6010",
            "100003",
            "   ",
            "5715",
            "Example Trust",
            "Primary",
            "Primary",
            "Not applicable",
            "Urban major conurbation",
            280,
            16,
            11.0,
            1.2,
            9.0,
            6.5,
            1400000,
            95000,
            1495000,
            720000,
            18000,
            190000,
            110000,
            1038000,
            45000,
            65000,
            38000,
            29000,
            36000,
            1410000,
            85000,
            35000,
        ]
    )

    central_services_sheet = workbook.create_sheet("Central Services")
    central_services_sheet.append(["Trust", "Total Income"])
    central_services_sheet.append(["Example Trust", 3000000])
    workbook.save(workbook_path)

    manifest_payload = {
        "normalization_contract_version": "school_financial_benchmarks.v1",
        "assets": [
            {
                "source_url": (
                    "https://financial-benchmarking-and-insights-tool.education.gov.uk/files/"
                    "AAR_2023-24_download.xlsx"
                ),
                "bronze_file_name": workbook_name,
                "downloaded_at": "2026-03-10T10:00:00+00:00",
                "sha256": "sha-finance",
                "academic_year": "2023/24",
                "sheet_names": ["Index", "Academies", "Central Services"],
                "sheet_row_counts": {
                    "Index": 1,
                    "Academies": 4,
                    "Central Services": 1,
                },
            }
        ],
    }
    (context.bronze_source_path / BRONZE_MANIFEST_FILE_NAME).write_text(
        json.dumps(manifest_payload),
        encoding="utf-8",
    )


def test_school_financial_benchmarks_pipeline_stage_and_promote_are_idempotent(
    engine: Engine,
    tmp_path: Path,
) -> None:
    bronze_root = tmp_path / "bronze"
    _seed_schools(engine)

    pipeline = SchoolFinancialBenchmarksPipeline(
        engine=engine,
        workbook_urls=(
            "https://financial-benchmarking-and-insights-tool.education.gov.uk/files/"
            "AAR_2023-24_download.xlsx",
        ),
    )

    first_context = _context(bronze_root)
    _write_manifest_and_workbook(first_context)
    _insert_run_row(engine, first_context)

    first_stage = pipeline.stage(first_context)
    assert first_stage.staged_rows == 2
    assert first_stage.rejected_rows == 2

    first_promoted = pipeline.promote(first_context)
    assert first_promoted == 2

    with engine.connect() as connection:
        finance_rows = connection.execute(
            text(
                """
                SELECT COUNT(*)
                FROM school_financials_yearly
                WHERE urn IN ('100001', '100002')
                """
            )
        ).scalar_one()
        assert finance_rows == 2

        alpha = connection.execute(
            text(
                """
                SELECT
                    total_income_gbp,
                    total_expenditure_gbp,
                    income_per_pupil_gbp,
                    expenditure_per_pupil_gbp,
                    staff_costs_pct_of_expenditure,
                    teaching_staff_costs_per_pupil_gbp,
                    supply_staff_costs_pct_of_staff_costs,
                    revenue_reserve_per_pupil_gbp,
                    source_file_url
                FROM school_financials_yearly
                WHERE urn = '100001' AND academic_year = '2023/24'
                """
            )
        ).one()
        assert alpha[0] == pytest.approx(2400000.0)
        assert alpha[1] == pytest.approx(2280000.0)
        assert alpha[2] == pytest.approx(6000.0)
        assert alpha[3] == pytest.approx(5700.0)
        assert alpha[4] == pytest.approx(1700000.0 / 2280000.0)
        assert alpha[5] == pytest.approx(2875.0)
        assert alpha[6] == pytest.approx(30000.0 / 1700000.0)
        assert alpha[7] == pytest.approx(800.0)
        assert (
            alpha[8] == "https://financial-benchmarking-and-insights-tool.education.gov.uk/files/"
            "AAR_2023-24_download.xlsx"
        )

        beta = connection.execute(
            text(
                """
                SELECT
                    pupils_fte,
                    teachers_fte,
                    total_income_gbp,
                    income_per_pupil_gbp,
                    expenditure_per_pupil_gbp,
                    staff_costs_pct_of_expenditure,
                    revenue_reserve_per_pupil_gbp
                FROM school_financials_yearly
                WHERE urn = '100002' AND academic_year = '2023/24'
                """
            )
        ).one()
        assert beta[0] is None
        assert beta[1] is None
        assert beta[2] is None
        assert beta[3] is None
        assert beta[4] is None
        assert beta[5] is None
        assert beta[6] is None

    second_context = _context(bronze_root)
    _insert_run_row(engine, second_context)
    second_stage = pipeline.stage(second_context)
    second_promoted = pipeline.promote(second_context)

    assert second_stage.staged_rows == 2
    assert second_stage.rejected_rows == 2
    assert second_promoted == 2

    with engine.connect() as connection:
        finance_rows = connection.execute(
            text(
                """
                SELECT COUNT(*)
                FROM school_financials_yearly
                WHERE urn IN ('100001', '100002')
                """
            )
        ).scalar_one()
        assert finance_rows == 2
